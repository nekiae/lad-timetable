"""Выгрузка расписания в Excel.

Зачем отдельный формат, если есть HTML. HTML — то, что пересылают и смотрят.
Excel — то, в чём завуч работает: печатает, правит, раздаёт учителям, вклеивает
в отчёт. Пока расписание нельзя выгрузить в книгу, оно живёт рядом с работой
школы, а не внутри неё.

Четыре листа, каждый отвечает на свой вопрос:
  «Классы»   — что у 7А в среду (сетка класс × слот, то же, что висит в коридоре);
  «Учителя»  — где Иванова в четверг третьим уроком (её личное расписание);
  «Кабинеты» — свободен ли спортзал во вторник вторым;
  «Проверка» — метрики и список нарушений, чтобы цифры можно было перепроверить.

Модуль ничего не знает про солвер: на вход — те же уроки, что и у рендера
и валидатора. Значит выгрузить можно и наше расписание, и реальное школьное.
"""

from collections import defaultdict
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .model import DayKind, Lesson, School
from .validate import check

DAY_NAMES = {1: "Понедельник", 2: "Вторник", 3: "Среда", 4: "Четверг",
             5: "Пятница", 6: "Суббота"}

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
DAY_FILL = PatternFill("solid", fgColor="D9E2F3")
EMPTY_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="center", horizontal="center")


def _header(sheet, titles: list[str], freeze: str = "B2") -> None:
    for column, title in enumerate(titles, start=1):
        cell = sheet.cell(row=1, column=column, value=title)
        cell.fill, cell.font, cell.alignment, cell.border = HEAD_FILL, HEAD_FONT, WRAP, BORDER
    sheet.freeze_panes = freeze
    sheet.row_dimensions[1].height = 28


def _slot_rows(school: School) -> list[tuple[int, int]]:
    """Все клетки сетки в порядке «день, номер урока»."""
    days = [d for d, kind in sorted(school.day_kinds.items()) if kind == DayKind.LESSONS]
    return [(day, period) for day in days for period in range(1, school.periods_per_day + 1)]


def _grid_sheet(sheet, school: School, columns: list[str], titles: list[str],
                content: dict, width: int = 26) -> None:
    """Общая раскладка: строки — слоты, столбцы — классы/учителя/кабинеты."""
    _header(sheet, ["День", "Урок"] + titles)
    sheet.column_dimensions["A"].width = 14
    sheet.column_dimensions["B"].width = 6
    for n in range(len(columns)):
        sheet.column_dimensions[get_column_letter(n + 3)].width = width

    row = 2
    previous_day = None
    for day, period in _slot_rows(school):
        day_cell = sheet.cell(row=row, column=1,
                              value=DAY_NAMES.get(day, str(day)) if day != previous_day else "")
        day_cell.fill = DAY_FILL
        day_cell.alignment = Alignment(vertical="center")
        day_cell.border = BORDER
        previous_day = day

        period_cell = sheet.cell(row=row, column=2, value=period)
        period_cell.alignment, period_cell.border = WRAP, BORDER

        for n, key in enumerate(columns):
            text = content.get((key, day, period), "")
            cell = sheet.cell(row=row, column=n + 3, value=text)
            cell.alignment, cell.border = WRAP, BORDER
            if not text:
                cell.fill = EMPTY_FILL
        sheet.row_dimensions[row].height = 32
        row += 1


def build_workbook(school: School, lessons: list[Lesson], anonymize: bool = False) -> Workbook:
    """Собрать книгу. anonymize — обезличить ФИО для показа третьим лицам (§8.4)."""
    subjects = {s.id: s.name for s in school.subjects}
    if anonymize:
        teachers = {t.id: f"Учитель {n + 1}" for n, t in enumerate(school.teachers)}
    else:
        teachers = {t.id: t.name for t in school.teachers}

    book = Workbook()

    # --- лист 1: по классам
    sheet = book.active
    sheet.title = "Классы"
    by_class: dict = defaultdict(list)
    for lesson in lessons:
        group = school.group(lesson.group_id)
        part = f" (гр. {group.part})" if group.part else ""
        text = subjects.get(lesson.subject_id, "?") + part
        text += "\n" + teachers.get(lesson.teacher_id, "")
        if lesson.room_id:
            text += f" · {lesson.room_id}"
        for class_id in group.class_ids:
            by_class[class_id, lesson.slot.day, lesson.slot.period].append(text)
    class_ids = [c.id for c in school.classes]
    _grid_sheet(sheet, school, class_ids, [c.name for c in school.classes],
                {k: "\n".join(v) for k, v in by_class.items()})

    # --- лист 2: по учителям
    sheet = book.create_sheet("Учителя")
    by_teacher: dict = defaultdict(list)
    for lesson in lessons:
        group = school.group(lesson.group_id)
        classes = "/".join(group.class_ids) + (f" (гр. {group.part})" if group.part else "")
        text = f"{classes}\n{subjects.get(lesson.subject_id, '?')}"
        if lesson.room_id:
            text += f" · {lesson.room_id}"
        by_teacher[lesson.teacher_id, lesson.slot.day, lesson.slot.period].append(text)
    teacher_ids = [t.id for t in school.teachers]
    _grid_sheet(sheet, school, teacher_ids, [teachers[t] for t in teacher_ids],
                {k: "\n".join(v) for k, v in by_teacher.items()}, width=22)

    # --- лист 3: по кабинетам
    sheet = book.create_sheet("Кабинеты")
    by_room: dict = defaultdict(list)
    for lesson in lessons:
        if not lesson.room_id:
            continue
        group = school.group(lesson.group_id)
        classes = "/".join(group.class_ids)
        by_room[lesson.room_id, lesson.slot.day, lesson.slot.period].append(
            f"{classes}\n{subjects.get(lesson.subject_id, '?')}")
    room_ids = [r.id for r in school.rooms]
    _grid_sheet(sheet, school, room_ids, room_ids,
                {k: "\n".join(v) for k, v in by_room.items()}, width=20)

    # --- лист 4: проверка
    sheet = book.create_sheet("Проверка")
    report = check(school, lessons)
    _header(sheet, ["Показатель", "Значение"], freeze="A2")
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 16
    row = 2
    for label, value in report.summary().items():
        sheet.cell(row=row, column=1, value=label).border = BORDER
        cell = sheet.cell(row=row, column=2, value=value)
        cell.border, cell.alignment = BORDER, Alignment(horizontal="center")
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="Нарушения").font = Font(bold=True)
    row += 1
    if report.violations:
        for column, title in enumerate(["Правило", "Что не так", "Где"], start=1):
            cell = sheet.cell(row=row, column=column, value=title)
            cell.fill, cell.font, cell.border = HEAD_FILL, HEAD_FONT, BORDER
        sheet.column_dimensions["C"].width = 22
        row += 1
        for violation in report.violations:
            sheet.cell(row=row, column=1, value=violation.rule).border = BORDER
            sheet.cell(row=row, column=2, value=violation.what).border = BORDER
            sheet.cell(row=row, column=3, value=violation.where).border = BORDER
            row += 1
    else:
        sheet.cell(row=row, column=1, value="Нарушений не найдено")

    return book


def to_bytes(school: School, lessons: list[Lesson], anonymize: bool = False) -> bytes:
    """Книга как байты — для кнопки скачивания."""
    buffer = BytesIO()
    build_workbook(school, lessons, anonymize).save(buffer)
    return buffer.getvalue()
